#author;R.Kunimoto, TAKENAKA co.
#coding:utf-8

import numpy as np
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill
import scipy.stats as stats
import math

def n_sigma(a,s,val):
    if val >= a+2*s or val <= a-2*s:
        return True
    else:
        return False

def SG_test(n,d,val,av,sd):
    t = stats.t.isf(d, n)
    significant_point = (n-1)*t/math.sqrt((n*(n-2))+(n*(t**2)))
    comp_point = (val-av)/sd
    if comp_point > significant_point:
        return True
    else:
        return False

def cossim(arr1,arr2):
    nor1 = np.linalg.norm(arr1)
    nor2 = np.linalg.norm(arr2)
    unko = np.dot(arr1,arr2)
    return unko/(nor1*nor2)

path = "C:\\Users\\1500570\\Documents\\R\\WS\\dataset_f2"
file = "170911_data_F2.csv"
wb = openpyxl.load_workbook(path+'\\170911_data_F2.xlsx', data_only = False)
ws = wb.get_sheet_by_name('170911_data_F2')

mama = re.compile(u"前月")

with open(path+"\\"+file, "r+") as f:
    reader = csv.DictReader(f)
    rf = reader.fieldnames
    rf3 = [] #NA許容範囲の属性名のインデクス配列
    p = 0

    reader2 = csv.reader(f)
    mat1 = [] #とりあえずcsvコピー

    for line in reader2:
        mat1.append(line)

    mat1 = np.array(mat1).T
    rf_term = mat1[0]
    mat1 = np.delete(mat1,0,0)

    mat3 = np.empty((0,len(mat1[0])))

    mm = 0
    for line in mat1:
        nona = np.where(line=="NA")[0]
        num = len(nona)
        tempmee = np.median(np.delete(line, np.where(line=="NA")).astype(np.float64))
        for i in nona:
            line[i] = tempmee
        if num <= 13:
            mat3 = np.vstack((mat3,np.array(line).astype(np.float64)))
            rf3.append(rf[mm])
        mm += 1
    mat3 = mat3.T
    print(mat3)

mat_result = []

i=0
while i < len(mat3):
    temp_row = []
    j = 0
    while j < len(mat3):
        temp_row.append(cossim(mat3[i], mat3[j]))
        j += 1
    mat_result.append(temp_row)
    i += 1
"""
print(mat_result)
np.savetxt(path+"\\teteteun.csv",mat_result,delimiter=",")
"""
row_avee = []
for unun in mat_result:
    row_avee.append(np.average(unun))

print(stats.shapiro(row_avee))

"""
#print(len(mat3))
#print(len(rf3))
ind1 = -1
for line in mat3:
    ind2 = 0
    temp = np.delete(line, np.where(line=="NA")).astype(np.float64)
    ind1 += 1
    if stats.shapiro(temp)[1] <= 0.05: #正規性検定
        ws.cell(row =1, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ffc040")
        continue
    else:
        ws.cell(row =1, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
        # 基本統計量算出
        av = np.average(temp)
        var = np.var(temp)
        sd = np.std(temp)
        #こっから各列の中身を見ていく
        maxabs = max([abs(max(temp)),abs(min(temp))])
        #print(max(temp), min(temp), line, temp)
        sigma = [0 for i in range(58)]
        sg = [0 for i in range(58)]

        for ii in line: #σ基準検出
            if ii != "NA":
                iii = ii.astype(np.float64)
                if n_sigma(av, sd, iii) == True:
                    sigma[ind2] = 1
            ind2 += 1

        key = True #スミルノフ・グラブス検定
        while key == True:
            maxabs = max([abs(max(temp)),abs(min(temp))])
            print(maxabs, max(temp), min(temp), line)
            indexes = [i for i, x in enumerate(line) if (x == str(maxabs) or x == str(-maxabs))]
            #print(indexes)
            if SG_test(len(temp), 0.05, maxabs, av, sd) == True:
                #print(rf3[ind1], ind2)
                #print(rf[rf3[ind1]], ind2)
                for i in indexes:
                    sg[i] = 1
                temp = np.delete(temp, np.where(temp==maxabs)).astype(np.float64)
                temp = np.delete(temp, np.where(temp==-maxabs)).astype(np.float64)
            else:
                key = False

        for i in range(58):
            if sigma[i] == 1:
                if sg == 1:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="9932cc")
                    print(rf3[ind1], i)
                    print(rf[rf3[ind1]], i)
                else:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
            else:
                if sg[i] == 1:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="0000ff")
                    #print(rf3[ind1], i)
                    #print(rf[rf3[ind1]], i)
                else:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ffffff")

wb.save(filename = path+'\\test_plus.xlsx')
"""
print("FINISHED")
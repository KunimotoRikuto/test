#author;R.Kunimoto, TAKENAKA co.
#coding:utf-8

import numpy as np
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill
import scipy.stats as stats
import math

def n_sigma(a,s,val):
    if val >= a+2*s or val <= a-2*s:
        return True
    else:
        return False

def SG_test(n,d,val,av,sd):
    t = stats.t.isf(d, n)
    significant_point = (n-1)*t/math.sqrt((n*(n-2))+(n*(t**2)))
    comp_point = (val-av)/sd
    if comp_point > significant_point:
        return True
    else:
        return False

path = "C:\\Users\\1500570\\Documents\\R\\WS\\dataset_f2"
file = "170911_data_F2.csv"
wb = openpyxl.load_workbook(path+'\\170911_data_F2.xlsx', data_only = False)
ws = wb.get_sheet_by_name('170911_data_F2')

mama = re.compile(u"前月")

with open(path+"\\"+file, "r+") as f:
    reader = csv.DictReader(f)
    rf = reader.fieldnames
    rf2 = [] #前月属性のインデクス配列
    rf3 = [] #NA許容範囲の属性名のインデクス配列
    p = 0
    for ff in rf:
        if mama.search(ff) != None:
            rf2.append(p)
        p += 1

    reader2 = csv.reader(f)
    mat1 = []
    for line in reader2:
        mat1.append(line)
    mat1 = np.array(mat1).T

    mat2 = np.empty((0,len(mat1[0])))
    for i in rf2:
        mat2 = np.vstack((mat2, mat1[i]))
    print(rf2)

    mat3 = np.empty((0,len(mat1[0])))
    mm = 0
    for line in mat2:
        num = len(np.where(line=="NA")[0])
        if num <= 13:
            mat3 = np.vstack((mat3,line))
            rf3.append(rf2[mm])
        mm += 1
    """
    # mat4 = mat3[1].astype(np.float64)
    mat4 = mat3[3]
    mat5 = np.delete(mat4, np.where(mat4=="NA"))
    print(mat4)
    print(mat5)
    # print(str(np.std(mat4)))
    print(rf[rf3[3]])
    """

#print(len(mat3))
#print(len(rf3))
ind1 = -1
for line in mat3:
    ind2 = 0
    temp = np.delete(line, np.where(line=="NA")).astype(np.float64)
    ind1 += 1
    if stats.shapiro(temp)[1] <= 0.05: #正規性検定
        ws.cell(row =1, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ffc040")
        continue
    else:
        ws.cell(row =1, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
        # 基本統計量算出
        av = np.average(temp)
        var = np.var(temp)
        sd = np.std(temp)
        ws.cell(row =ind2+2, column =1).fill = PatternFill(patternType="solid", fgColor="ffc040")
        #こっから各列の中身を見ていく
        
        for ii in line: #σ基準検出
            if ii != "NA":
                iii = ii.astype(np.float64)
                if n_sigma(av, sd, iii) == True:
                    ws.cell(row =ind2+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
                    print(rf3[ind1], ind2)
                    print(rf[rf3[ind1]], ind2)
            ind2 += 1
        

        for ii in line: #スミルノフ・グラブス検定
            maxabs = max([abs(max(temp)),abs(min(temp))])
            if ii != "NA":
                iii = ii.astype(np.float64)
                if SG_test(len(temp), 0.05, iii, av, sd) == True:
                    ws.cell(row =ind2+2, column =1).fill = PatternFill(patternType="solid", fgColor="ffc040")
                else:
                    ws.cell(row =ind2+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
                print(rf3[ind1], ind2)
                print(rf[rf3[ind1]], ind2)
            ind2 += 1

wb.save(filename = path+'\\test_sigma.xlsx')

print("FINISHED")
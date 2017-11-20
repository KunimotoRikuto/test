#author;R.Kunimoto, TAKENAKA co.
#coding:utf-8

import numpy as np
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill
import scipy.stats as stats
import math
import matplotlib.pyplot as plt
from sklearn import cluster
from collections import Counter
import random

def n_sigma(a,s,val):
    if val >= a+2*s or val <= a-2*s:
        return True
    else:
        return False

def SG_test(n,d,val,av,sd):
    t = stats.t.isf(d, n)
    significant_point = (n-1)*t/math.sqrt((n*(n-2))+(n*(t**2)))
    comp_point = (val-av)/sd
    if comp_point > significant_point:
        return True
    else:
        return False

def generate_random_color():
    return '{:X}{:X}{:X}'.format(*[random.randint(0, 255) for _ in range(3)])

colorbar = ["008b8b","008080","2f4f4f","006400","008000","228b22","2e8b57","3cb371","66cdaa","8fbc8f","7fffd4","98fb98"]

path = "C:\\Users\\1500570\\Documents\\R\\WS\\dataset_f2"
file = "170911_data_F2.csv"
wb = openpyxl.load_workbook(path+'\\170911_data_F2.xlsx', data_only = False)
ws = wb.get_sheet_by_name('170911_data_F2')

mama = re.compile(u"前月")

with open(path+"\\"+file, "r+") as f:
    reader = csv.DictReader(f)
    rf = reader.fieldnames
    rf2 = [] #前月属性のインデクス配列
    rf3 = [] #NA許容範囲の属性名のインデクス配列
    p = 0
    for ff in rf:
        if mama.search(ff) != None:
            rf2.append(p)
        p += 1

    reader2 = csv.reader(f)
    mat1 = []
    for line in reader2:
        mat1.append(line)
    mat1 = np.array(mat1).T

    mat2 = np.empty((0,len(mat1[0])))
    for i in rf2:
        mat2 = np.vstack((mat2, mat1[i]))
    print(rf2)

    mat3 = np.empty((0,len(mat1[0])))
    mm = 0
    for line in mat2:
        num = len(np.where(line=="NA")[0])
        if num <= 13:
            mat3 = np.vstack((mat3,line))
            rf3.append(rf2[mm])
        mm += 1
    """
    # mat4 = mat3[1].astype(np.float64)
    mat4 = mat3[3]
    mat5 = np.delete(mat4, np.where(mat4=="NA"))
    print(mat4)
    print(mat5)
    # print(str(np.std(mat4)))
    print(rf[rf3[3]])
    """

#print(len(mat3))
#print(len(rf3))
ind1 = -1
row_false = []
row_true = []
pltnum = 0
for line in mat3:
    ind2 = 0
    temp = np.delete(line, np.where(line=="NA")).astype(np.float64)
    plt.hist(temp)
    plt.xlim([max(temp),min(temp)])
    plt.savefig(path+"\\"+rf[rf3[ind1+1]]+".png")
    plt.close()
    pltnum += 1
    ind1 += 1
    if stats.shapiro(temp)[1] <= 0.05: #正規性検定
        ws.cell(row =1, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ffc040")
        row_false.append(rf[rf3[ind1]])
        temp2 = np.array([[0]])
        for i in temp:
            temp2 = np.vstack((temp2,i))
        temp2 = np.delete(temp2,0,0)
        print(temp2)
        #kmeans = cluster.KMeans(n_clusters=3).fit(temp2)
        meanshift = cluster.MeanShift(seeds=temp2).fit(temp2)
        #kmm = kmeans.labels_
        mam = meanshift.labels_
        print (mam)
        break
        """
        counter = Counter(mam)
        cc = 0
        for cnt in counter:
            cc += 1
        trwi = []
        for i in range(cc):
            trwi.append(generate_random_color())
        #print(trwi)
        """
        pika = 0
        ink = 0
        #print(line, temp)
        while pika < 58:
            #print(line[pika],print(temp2[ink][0]))
            if line[pika] != "NA":
                #print(mam[ink])
                ws.cell(row =pika+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor=colorbar[mam[ink]])
                pika += 1
                ink += 1
            else:
                #print(line[pika])
                pika += 1
    else:
        ws.cell(row =1, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
        row_true.append(rf[rf3[ind1]])
        # 基本統計量算出
        av = np.average(temp)
        var = np.var(temp)
        sd = np.std(temp)
        #こっから各列の中身を見ていく
        maxabs = max([abs(max(temp)),abs(min(temp))])
        #print(max(temp), min(temp), line, temp)
        sigma = [0 for i in range(58)]
        sg = [0 for i in range(58)]

        for ii in line: #σ基準検出
            if ii != "NA":
                iii = ii.astype(np.float64)
                if n_sigma(av, sd, iii) == True:
                    sigma[ind2] = 1
            ind2 += 1

        key = True #スミルノフ・グラブス検定
        while key == True:
            maxabs = max([abs(max(temp)),abs(min(temp))])
            print(maxabs, max(temp), min(temp), line)
            indexes = [i for i, x in enumerate(line) if (x == str(maxabs) or x == str(-maxabs))]
            #print(indexes)
            if SG_test(len(temp), 0.05, maxabs, av, sd) == True:
                #print(rf3[ind1], ind2)
                #print(rf[rf3[ind1]], ind2)
                for i in indexes:
                    sg[i] = 1
                temp = np.delete(temp, np.where(temp==maxabs)).astype(np.float64)
                temp = np.delete(temp, np.where(temp==-maxabs)).astype(np.float64)
            else:
                key = False

        for i in range(58):
            if sigma[i] == 1:
                if sg == 1:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="9932cc")
                    print(rf3[ind1], i)
                    print(rf[rf3[ind1]], i)
                else:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ff0000")
            else:
                if sg[i] == 1:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="0000ff")
                    #print(rf3[ind1], i)
                    #print(rf[rf3[ind1]], i)
                else:
                    ws.cell(row =i+2, column =rf3[ind1]+1).fill = PatternFill(patternType="solid", fgColor="ffffff")

wb.save(filename = path+'\\test_plus.xlsx')

"""
np.savetxt(path+"\\row_false.csv",row_false,delimiter=",")
np.savetxt(path+"\\row_true.csv",row_false,delimiter=",")
"""

print("FINISHED")
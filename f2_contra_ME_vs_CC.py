#author;R.Kunimoto, TAKENAKA co.
#coding:utf-8

import numpy as np
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill
import scipy.stats as stats
import math
import matplotlib.pyplot as plt
from sklearn import cluster
from collections import Counter
import random
from pygments.lexers.csound import newline


markers = (".",",","o","v","^","<" ,">","1","2","3")
colorlist = ["r", "g", "b", "c", "m", "y", "k", "r", "g", "b", "c", "m", "y", "k"]

def n_sigma(a,s,val):
    if val >= a+2*s or val <= a-2*s:
        return True
    else:
        return False

def SG_test(n,d,val,av,sd):
    t = stats.t.isf(d, n)
    significant_point = (n-1)*t/math.sqrt((n*(n-2))+(n*(t**2)))
    comp_point = (val-av)/sd
    if comp_point > significant_point:
        return True
    else:
        return False

def generate_random_color():
    return '{:X}{:X}{:X}'.format(*[random.randint(0, 255) for _ in range(3)])

colorbar = ["008b8b","008080","2f4f4f","006400","008000","228b22","2e8b57","3cb371","66cdaa","8fbc8f","7fffd4","98fb98"]

path = "C:\\Users\\1500570\\Documents\\R\\WS\\dataset_f2"
file = "170911_data_F2.csv"
wb_m = openpyxl.load_workbook(path+'\\170911_data_F2.xlsx', data_only = False)
wb_c = openpyxl.load_workbook(path+'\\comments.xlsx', data_only = False)
ws_m = wb_m.get_sheet_by_name('170911_data_F2')
ws_c = wb_c.get_sheet_by_name('Sheet1')

i = 2
commented_field_index = []
mat_scope = []
mat_commented_col = []
mat_commented_raws_index = []

while i <= 1156:
    j = 2
    temp = []
    temp2= []
    flag = False
    na_c = 0
    while j <= 59:
        if ws_c.cell(row =j, column =i).value != "NA":
            temp.append(float(ws_c.cell(row =j, column =i).value))
        else:
            temp.append(ws_c.cell(row =j, column =i).value)
            na_c += 1
        if ws_c.cell(row =j, column =i).fill.start_color.index != "00000000":
            temp2.append(j)
            flag = True
        j += 1
    if flag == True and na_c <= 13:
        #print(temp)
        mat_scope.append(temp)
        mat_commented_raws_index.append(temp2)
        commented_field_index.append(i)
    i += 1

unko = []

ind1 = -1
row_false = []
row_true = []
ep = 1
line_num = 0
mat_scope = np.array(mat_scope)
print("matscope",len(mat_scope))

for line in mat_scope:
    ind2 = 0
    temp = np.delete(line, np.where(line=="NA")).astype(np.float64)
    temp2 = np.array([[0]])
    for i in temp:
        temp2 = np.vstack((temp2,i))
    pika = 0
    pika_temp = 0
    if stats.shapiro(temp2)[1] <= 0.05: #正規性検定
        print("cluster")
        meanshift = cluster.MeanShift(seeds=temp2,min_bin_freq=10).fit(temp2)
        #kmm = kmeans.labels_
        mam = meanshift.labels_
        while pika < len(temp):
            #print(line[pika],print(temp2[ink][0]))
            if ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value != "NA":
                plt.subplot(2,1,1)
                plt.grid(True)
                if ws_c.cell(row =pika+2, column =commented_field_index[line_num]).fill.start_color.index == "00000000":
                    plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c="black")
                else:
                    plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c="red")
                plt.subplot(2,1,2)
                plt.grid(True)
                plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c=colorlist[mam[pika+1]], marker=markers[mam[pika+1]])
                pika_temp += 1
            else:
                mam = np.insert(mam,pika,0)
            pika += 1
        plt.savefig(path+"\\"+str(ws_c.cell(row =1, column =commented_field_index[line_num]).value)+".png")
        unko.append(str(ws_c.cell(row =1, column =commented_field_index[line_num]).value))
        plt.close()
        line_num += 1
        print(line_num,"th")
    else:
        print("gauss")
        # 基本統計量算出
        av = np.average(temp)
        var = np.var(temp)
        sd = np.std(temp)
        #こっから各列の中身を見ていく
        maxabs = max([abs(max(temp)),abs(min(temp))])
        #print(max(temp), min(temp), line, temp)
        sigma = [0 for i in range(58)]
        sg = [0 for i in range(58)]

        for ii in line: #σ基準検出
            if ii != "NA":
                iii = ii.astype(np.float64)
                if n_sigma(av, sd, iii) == True:
                    sigma[ind2] = 1
            ind2 += 1

        key = True #スミルノフ・グラブス検定
        while key == True:
            maxabs = max([abs(max(temp)),abs(min(temp))])
            print(maxabs, max(temp), min(temp), line)
            indexes = [i for i, x in enumerate(line) if (x == str(maxabs) or x == str(-maxabs))]
            #print(indexes)
            if SG_test(len(temp), 0.05, maxabs, av, sd) == True:
                #print(rf3[ind1], ind2)
                #print(rf[rf3[ind1]], ind2)
                for i in indexes:
                    sg[i] = 1
                temp = np.delete(temp, np.where(temp==maxabs)).astype(np.float64)
                temp = np.delete(temp, np.where(temp==-maxabs)).astype(np.float64)
            else:
                key = False
        while pika < len(temp):
            if ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value != "NA":
                plt.subplot(2,1,1)
                plt.grid(True)
                if ws_c.cell(row =pika+2, column =commented_field_index[line_num]).fill.start_color.index == "00000000":
                    plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c="black")
                else:
                    plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c="red")
                plt.subplot(2,1,2)
                plt.grid(True)
                if sigma[pika] == 1 or sg[pika] == 1:
                    plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c="red")
                else:
                    plt.scatter(pika,float(ws_c.cell(row =pika+2, column =commented_field_index[line_num]).value),c="gray",marker=".")
                pika_temp += 1
            else:
                np.insert(mam,pika,0)
            pika += 1
        plt.savefig(path+"\\"+str(ws_c.cell(row =1, column =commented_field_index[line_num]).value)+".png")
        plt.close()
        line_num += 1
        print(line_num,"th")

print(mat_scope)
print(commented_field_index)

print(unko)
print("finished")
